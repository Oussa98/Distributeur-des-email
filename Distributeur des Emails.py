#author : Oussama BRAHITI

################## IMPORTATION des MODULES et des PACKAGES ####################

# Importation des modules reliés aux GUI
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QLabel, QLineEdit, QComboBox, QPlainTextEdit, QRadioButton, QFrame, QCheckBox, QMessageBox, QProgressBar
from PyQt5.QtCore import QRect, Qt, QSize
from PyQt5.QtGui import QFont, QIcon
from sys import argv, exit

# # # # # # #Importation des modules de la logic:
#openpyxl pour lire les fichier .xlsx
from openpyxl import load_workbook 
from openpyxl.utils.exceptions import InvalidFileException
#smtplib pour envoyer les email 
from smtplib import SMTP
from smtplib import SMTPAuthenticationError
from smtplib import SMTPSenderRefused
from smtplib import SMTPRecipientsRefused
# email pour formater les email (les decoder en utf-8 ... ect)
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
# datetime pour avoir le temps et la date du jour 
import datetime
mainent = datetime.datetime.now() # pour avoir la date et le temps de maintenant

############################## les utilitées #############################

class DnD(QLabel): #subclass la class de QLabel pour autoriser l'effet Trainez puis Relacher

    file_path = ""
    file = None

    def __init__(self, parent = None):
 
        super().__init__(parent)
        self.setAcceptDrops(True)


    def dragEnterEvent(self, event): # quand le fichier traîné entre dans le QLabel
        if event.mimeData().hasUrls:
            event.accept()
        else:
            event.ignore()

    def dragMoveEvent(self, event): # quand le fichier traîné bouge dans le QLabel
        if event.mimeData().hasUrls():
            event.accept()
        else:
            event.ignore()

    def dropEvent(self, event): # quand le fichier traîné est relaché
        if event.mimeData().hasUrls():
            event.accept()
            self.file_path = event.mimeData().urls()[0].toLocalFile()
            self.file_path.encode('utf-8')
            try:
                self.parent().ch7al.setText('')
                DnD.file = load_workbook(self.file_path)
                
                feuilles = DnD.file.sheetnames
                self.parent().choisir_feuille.clear()
                self.parent().charger_feuille(feuilles)
                self.parent().trainez.setText("xlsx Chargé")
                self.parent().my_button.setEnabled(True)
                self.parent().progres.setValue(0)
                ff = DnD.file[self.parent().choisir_feuille.currentText()]
                self.parent().ch7al.setText(str(ff.max_row - 4))
                DnD.file_path = self.file_path
                # DnD.file = 
            except InvalidFileException:
                self.parent().ch7al.setText('')
                self.parent().progres.setValue(0)
                self.parent().choisir_feuille.clear()
                self.parent().choisir_feuille.setEnabled(False)
                self.parent().my_button.setEnabled(False)
                self.parent().trainez.setText("OPS !! Le Fichier Que Vous Avez Trainez \n n'est pas un Fichier .XLSX \n Veuillez essayer de l'ouvrir sois avec \n Microsoft Office - Excel ou bien \n avec Libre Office - Calc ")
        else:
            event.ignore()

############################ l'application générale
class App(QWidget): #l'application en grosomodo

    #class variables
    infos = [2, 3]
    safe = True

    
    def __init__(self): #fonction initialization de la class
        super().__init__()
        self.title = "Distrubiteur des Emails"        
        self.width = 640
        self.height = 465

        self.win() # appeler la fonction win


    def win(self): # pour fabriquer la fenêtre

        self.setWindowTitle(self.title)
        self.setWindowIcon(QIcon('mail_32.ico')) 
        self.setMinimumSize(QSize(640, 465))
        self.setMaximumSize(QSize(640, 465))
        self.resize(640, 465)

        self.contenu() # appler la fonction contenu

        self.show() # pour voir les composant de win et contenu

    def contenu(self):
        # || c'est le button pour envoyer les emails ||
        self.my_button = QPushButton("Envoyer !", self)
        self.my_button.setGeometry(QRect(555, 405, 70, 50))
        self.my_button.setObjectName("envoyer")
        self.my_button.setEnabled(False)
        self.my_button.clicked.connect(self.on_click)
        
        #  || c'est le button pour afficher a propos ||
        self.apropos_button = QPushButton("à propos", self)
        self.apropos_button.setGeometry(QRect(460, 405, 75, 25))
        self.apropos_button.setObjectName("apropos")
        self.apropos_button.clicked.connect(self.apropos)
        
        # || c'est le button pour afficher les limitations ||
        self.limitation_button = QPushButton("Limitations", self)
        self.limitation_button.setGeometry(QRect(460, 430, 75, 25))
        self.limitation_button.setObjectName("limit")
        self.limitation_button.clicked.connect(self.limitation)
        
        # || c'est une barre de progres (les emails envoyés) ||
        self.progres = QProgressBar(self)
        self.progres.setGeometry(QRect(10, 425, 400, 30))

        # || c'est pour l'effet Drag and Drop (Trainez puis Relacher) ||
        self.trainez = DnD(self)
        self.trainez.setAcceptDrops(True)
        self.trainez.setText("Trainez votre Fichier .xlsx ici")
        
        font = QFont()
        font.setPointSize(10)
        self.trainez.setFont(font)
        
        self.trainez.setAlignment(Qt.AlignCenter)
        self.trainez.setGeometry(QRect(10, 10, 258, 212))
        self.trainez.setStyleSheet("QLabel{border: 4px dashed #aaa}")
        self.trainez.setObjectName("leffet_dnd")
        
        # || c'est pour savoir combien de lignes ||
        self.ch7al = QLabel(self)
        self.ch7al.setGeometry(QRect(285, 390, 30, 30))
        
        # || c'est pour Introduire un l'email ||
        self.avoir_email = QLineEdit(self)
        self.avoir_email.setGeometry(QRect(10, 240, 284, 31))
        self.avoir_email.setObjectName("pour_email")

        font = QFont()
        font.setPointSize(12)
        self.avoir_email.setFont(font)
        
        self.avoir_email.setText("")
        self.avoir_email.setPlaceholderText("Email")

        # || c'est pour Introduire le mot de passe pour l'email ||
        self.avoir_mot_de_passe = QLineEdit(self)
        self.avoir_mot_de_passe.setGeometry(QRect(10, 280, 284, 31))
        self.avoir_mot_de_passe.setObjectName("pour_mot_de_passe")

        font = QFont()
        font.setPointSize(12)
        self.avoir_mot_de_passe.setFont(font)
        
        self.avoir_mot_de_passe.setText("")
        self.avoir_mot_de_passe.setPlaceholderText("Mot de passe")
        self.avoir_mot_de_passe.setEchoMode(QLineEdit.Password)

        # || c'est pour le choix du serveur SMTP ||
        self.choisir_serveur = QComboBox(self)
        self.choisir_serveur.setGeometry(QRect(20, 325, 250, 31))
        self.choisir_serveur.setObjectName("Serveur")

        font = QFont()
        font.setPointSize(12)
        self.choisir_serveur.setFont(font)

        self.choisir_serveur.addItem("Choisissez Votre Serveur SMTP")
        self.choisir_serveur.addItem("Gmail")
        self.choisir_serveur.addItem("Outlook")
        self.choisir_serveur.addItem("Yahoo")

        # || c'est pour le choix de la feuille du travail ||
        text_feuille = QLabel("Choisissez Votre xlsx feuille", self)
        text_feuille.setGeometry(QRect(20, 360, 250, 31))
        font = QFont()
        font.setPointSize(12)
        text_feuille.setFont(font)
        # || c'est pour le choix de la feuille du travail ||
        self.choisir_feuille = QComboBox(self)
        self.choisir_feuille.setGeometry(QRect(20, 390, 250, 31))
        self.choisir_feuille.setObjectName("Feuilles")
        self.choisir_feuille.setEnabled(False)
        self.choisir_feuille.activated.connect(self.charger_ch7al)
        font = QFont()
        font.setPointSize(12)
        self.choisir_feuille.setFont(font)

        # || c'est pour les notes supplimentaires ||
        self.notabene = QPlainTextEdit(self)
        self.notabene.setGeometry(QRect(310, 20, 300, 90))
        self.notabene.setObjectName("pourconsultation")

        font = QFont()
        font.setPointSize(12)
        self.notabene.setFont(font)

        self.notabene.setPlaceholderText("Ajouter des Notes; commme pour \n la date de la Consultation")
        self.notabene.setAcceptDrops(False)

        # || c'est pour choisir les information à envoyer au étudiant ||
        self.les_colonnes = QFrame(self) # construire une FRAME special pour les CheckBox
        self.les_colonnes.setGeometry(QRect(310, 120, 300, 270))
        ## self.les_colonnes.setStyleSheet("background-color: rgb(255, 255, 255);")
        
        self.titre_des_checks = QLabel("Choisir les informations à envoyer a \n vos étudiants : ", self.les_colonnes)
        font = QFont()
        font.setPointSize(10)
        self.titre_des_checks.setFont(font)
        ## self.les_colonnes.setGeometry(QRect(0, 0, 200, 60))


        self.examen = QCheckBox("examen", self.les_colonnes)
        self.examen.setGeometry(QRect(5, 30, 150, 30))
        
        self.rattrapage = QCheckBox("Rattrapage", self.les_colonnes)
        self.rattrapage.setGeometry(QRect(5, 60, 150, 30))
        
        self.exam_de_remp = QCheckBox("examen de remplaçement", self.les_colonnes)
        self.exam_de_remp.setGeometry(QRect(5, 90, 150, 30))
        
        self.assiduity = QCheckBox("L'assiduité", self.les_colonnes)
        self.assiduity.setGeometry(QRect(5, 120, 150, 30))
        
        self.presence = QCheckBox("La présence", self.les_colonnes)
        self.presence.setGeometry(QRect(5, 150, 150, 30))
        
        self.test_01 = QCheckBox("Test 01", self.les_colonnes)
        self.test_01.setGeometry(QRect(5, 180, 150, 30))
        
        self.test_02 = QCheckBox("Test 02", self.les_colonnes)
        self.test_02.setGeometry(QRect(75, 180, 150, 30))
        
        self.test_03 = QCheckBox("Test 03", self.les_colonnes)
        self.test_03.setGeometry(QRect(145, 180, 150, 30))
        
        self.test_04 = QCheckBox("Test 04", self.les_colonnes)
        self.test_04.setGeometry(QRect(215, 180, 150, 30))
        
        self.moy_des_tests = QCheckBox("Moyenne des tests", self.les_colonnes)
        self.moy_des_tests.setGeometry(QRect(5, 210, 150, 30))
        
        self.mean = QCheckBox("La moyenne génerale", self.les_colonnes)
        self.mean.setGeometry(QRect(5, 240, 150, 30))
        
    def combo_serveur(self): # pour avoir le serveur de l'email apartir du ComboBox
    
        if self.choisir_serveur.currentText() == "Gmail":
            self.server = "smtp.gmail.com"
        elif self.choisir_serveur.currentText() == "Yahoo":
            self.server = "smtp.mail.yahoo.com"
        elif self.choisir_serveur.currentText() == "Outlook":
            self.server = "smtp.office365.com"
    def charger_ch7al(self):
    
        file = DnD.file # avoir le chemin par DnD
        war9a = file[str(self.choisir_feuille.currentText())]
        self.ch7al.setText(str(war9a.max_row - 4))
        
    def charger_feuille(self, feuilles): # pour charger les SpreadSheets de l'excel (appel dans la DnD class)
    
        for feuil in feuilles:
            self.choisir_feuille.addItem(feuil)
        self.choisir_feuille.setEnabled(True)

    def combo_colonnes(self): # pour avoir l'etat des CheckBoxes (cocher ou non)

        if self.examen.isChecked() and 4 not in self.infos:
            self.infos.append(4)
        elif self.examen.isChecked() == False and 4 in self.infos:
            self.infos.remove(4)

        if self.rattrapage.isChecked() and 5 not in self.infos:
            self.infos.append(5)
        elif self.rattrapage.isChecked() == False and 5 in self.infos:
            self.infos.remove(5)

        if self.exam_de_remp.isChecked() and 6 not in self.infos:
            self.infos.append(6)
        elif self.exam_de_remp.isChecked() == False and 6 in self.infos:
            self.infos.remove(6)


        if self.assiduity.isChecked() and 7 not in self.infos:
            self.infos.append(7)
        elif self.assiduity.isChecked() == False and 7 in self.infos:
            self.infos.remove(7)

        if self.presence.isChecked() and 8 not in self.infos:
            self.infos.append(8)
        elif self.presence.isChecked() == False and 8 in self.infos:
            self.infos.remove(8)


        if self.test_01.isChecked() and 9 not in self.infos:
            self.infos.append(9)
        elif self.test_01.isChecked() == False and 9 in self.infos:
            self.infos.remove(9)

        if self.test_02.isChecked() and 10 not in self.infos:
            self.infos.append(10)
        elif self.test_02.isChecked() == False and 10 in self.infos:
            self.infos.remove(10)

        if self.test_03.isChecked() and 11 not in self.infos:
            self.infos.append(11)
        elif self.test_03.isChecked() == False and 11 in self.infos:
            self.infos.remove(11)

        if self.test_04.isChecked() and 12 not in self.infos:
            self.infos.append(12)
        elif self.test_04.isChecked() == False and 12 in self.infos:
            self.infos.remove(12)


        if self.moy_des_tests.isChecked() and 13 not in self.infos:
            self.infos.append(13)
        elif self.moy_des_tests.isChecked() == False and 13 in self.infos:
            self.infos.remove(13)

        if self.mean.isChecked() and 14 not in self.infos:
            self.infos.append(14)
        elif self.mean.isChecked() == False and 14 in self.infos:
            self.infos.remove(14)

        return self.infos

    def on_click(self): # quoi arrive til quand le bouton envoyer ! est cliqué
        self.progres.setValue(0)
        self.note = self.notabene.document().toPlainText()   
        self.combo_colonnes()
        self.themessages()
        if self.safe == True:
            try:
                self.lire(self.choisir_feuille.currentText())
            except :
                msg = QMessageBox()
                msg.setWindowTitle("Fichier Excel Corrempu|| Internet")
                msg.setText("Veuillez vérifier votre fichier excel ou votre connection internet")
                x = msg.exec_()
        else:
            pass
    def themessages(self): # c'est pour attraper des érreurs pour empêcher le crash du programme
        self.safe == True
        if self.infos == [2, 3] and self.note == "":
            self.safe = False
            msg = QMessageBox()
            msg.setWindowTitle("Information Necessaire")
            msg.setText("Veuillez choisir ce que vous voulez envoyer a vos étudiant \n en cochant les carrés \n ou bien \n en ecrivé une NotaBene")
            x = msg.exec_()
        if self.choisir_serveur.currentText() == "Choisissez Votre Serveur SMTP":
            self.safe = False
            msg = QMessageBox()
            msg.setWindowTitle("Information Necéssaire")
            msg.setText("Veuillez choisir fournisseur de votre email")
            x = msg.exec_()
        if ("@" and '.') not in self.avoir_email.text():
            self.safe = False
            msg = QMessageBox()
            msg.setWindowTitle("Information Necéssaire")
            msg.setText("Veuillez Entrez votre email correctement")
            x = msg.exec_()
        if self.avoir_mot_de_passe.text() == "":
            self.safe = False
            msg = QMessageBox()
            msg.setWindowTitle("Information Necéssaire")
            msg.setText("Veuillez Entrez votre mot de passe")
            x = msg.exec_()            
        
    def lire(self, feuil): # lire le fichier xlsx et choisir quoi choisir

        self.combo_serveur()
        file = DnD.file # avoir le chemin par DnD
        war9a = file[feuil] # charger la feuile selecter
        
        ROWS = war9a.max_row
        print("ligne {} dans l'excel".format(ROWS))
        therows = ROWS - 5
        nomdumodule = str(war9a.cell(row = 2, column = 2).value)
        departement = str(war9a.cell(row = 2, column = 4).value) + str(war9a.cell(row = 2, column = 5).value) + "\n"
        section = str(war9a.cell(row = 2, column = 6).value) + str(war9a.cell(row = 2, column = 7).value) + "\n"
        groupe = str(war9a.cell(row = 2, column = 8).value) + str(war9a.cell(row = 2, column = 9).value)
        
        self.note = self.notabene.document().toPlainText()           
        self.progres.setMaximum(therows+1)
        for ligne in range(5, ROWS +1):
            print("ligne : {}".format(ligne-4))
            
            # self.ch7al.setText(str(therows-5))
            etudiant = ""
            receveur = str(war9a.cell(row = ligne, column = 1).value) # l'email a qui il faut envoyer un email
            print(receveur)
            for colonne in self.infos:
                etudiant = etudiant + str(war9a.cell(row = 4, column = colonne).value) + str(war9a.cell(row = ligne, column = colonne).value) + "\n"

            try:  # Try et Except pour attraper des erreur et ne pas laisser le system crashé
                self.envoyer(self.server, etudiant, self.note, nomdumodule, departement, section, groupe, receveur)
                self.progres.setValue(ligne-4)

            except SMTPRecipientsRefused:
                msg = QMessageBox()
                msg.setWindowTitle("Information erroné")
                msg.setText("Impossible de se connecter a l'email de l'etudiant qui est  \n Veuillez vérifiez son compte")
                x = msg.exec_()
                continue

            except SMTPAuthenticationError:
                msg = QMessageBox()
                msg.setWindowTitle("Information erroné")
                msg.setText("""Impossible de se connecter a votre email \n Veuillez vérifiez votre compte; \n  ou bien email ou/et mot de passe \n et Veuillez verifiez le serveur SMTP de votre email \n ou bien autoriser les applications moins sécurisé d'utilizer votre email""")
                # msg.setOpenExternalLinks(True) # Failed xxxx Annuler
                x = msg.exec_()
                break

            except :
                msg = QMessageBox()
                msg.setWindowTitle("ERREUR//ERREUR//ERREUR")
                msg.setText("Erreur Inconnue :(  ")
                x = msg.exec_()

    def envoyer(self, server, etudiant, note, nomdumodule, departement, section, groupe, receveur): #envoyer les email
        msg = MIMEMultipart()

        msg['From'] = self.avoir_email.text()
        password = self.avoir_mot_de_passe.text()
        msg['To'] = receveur
        a = "\n ************************* \n "
        message = a + departement + section + groupe + a + etudiant + a + note + a # le corp de l'email
        msg['Subject'] = "UNIV_BOUIRA " + nomdumodule + " " + mainent.strftime("%d-%m-%Y") # le sujet de l'email

        msg.attach(MIMEText(message, 'plain')) # attacher le message a l'email (mode plain opposé au mode HTML)
        server = SMTP(server, 587) # commencer la connection avec le serveur choisis par le port 587
        
        server.ehlo() # dire Hello au serveur
        server.starttls() # commencer la connection en TLS
        server.ehlo() # dire Hello au serveur
        
        server.login(msg['From'], password) # Se Connecter au serveur SMTP

        server.sendmail(msg['From'], msg['To'], msg.as_string()) # enfin envoyer les emails
    
    def apropos(self): # quoi arrive til quand le bouton à propos est cliqué
        msg = QMessageBox()
        msg.setWindowTitle("Information sur ce programme")
        msg.setText(r"""c'est quoi ça:
        Ce programme à été réalisé par l'etudiant Oussama BRAHITI durant la quarantaine (le confinement du COVID-19) et comme réponse a une compétition organizée par :
        
                        l'université de Akli Mohand OULHADJ de BOUIRA - ALGERIE 
                        
pour quoi ce programme? :

    ""Cet outil est destiné aux enseignants pour communiquer avec leurs étudiants d'une manière professionelle et économique""
    
    C'est une humble solution pour résoudre le désordre de l'affichage des résultats et comme tentative d'économiser l'energie des étudiants qui font la navette chaque jour pour seulment voire leurs notes ou avoir des nouvelles.
    
Comment utilizer ce programme : 
    1) Completez le fichier excel accompagné avec ce programme.
    
    2) Trainez le fichier excel vers la boite en haut à gauche de la fenêtre, puis relachez
    
    3) Entrez vos coordonnées (Email et mot de passe ) 
        ( activer d'abbord l'option [autoriser les applications moins sécurisées])
    
    4) Choisissez votre serveur SMTP( le fournisseur de votre email )
    
    5) Choisissez votre feuille excel (SpreadSheet)
    
    6) Entrez une N.B si vous voulez (optionel)
    
    7) Cochez les information à envoyer a vos étudiants
    
    8) clickez le bouton Envoyer !""")
        x = msg.exec_()          
    
    def limitation(self): # quoi arrive til quand le bouton limitation est cliqué
        self.safe = False
        msg = QMessageBox() 
        msg.setWindowTitle("Limitation de ce programme")
        msg.setText("""Ce programme peut être parfais si ces limitations n'exsitaient pas
        
        1/ Ce programme ne contient pas un serveur special(ça côut de l'argent) Donc :
        
         -- les nombre d'emails qui peuvent être envoyés depuis un email depend du fournisseur:
        Google  ----> 500 par jour
        Microsoft(outlook, live, hotmail) ----> selon votre compte, environ 300 par jour
        Yahoo ----> 500 par jour
             
         -- la sécurité et confidnetialité : selon le service
         
        2/ il faut préparer l'email d'abbord; par l'activation de l'option (autoriser les applications moins sécurisées)
    chez Google c'est par la : "https://myaccount.google.com/lesssecureapps"
    chez Yahoo c'est par la : "https://login.yahoo.com/account/security#other-apps"
    chez Microsoft c'est compliqué, vous cherchez dans les paramètres.
        
        3/ vous avez seulment trois fournisseur d'email les plus célebres (Google, Yahoo, Microsoft)
        """)
        x = msg.exec_()  
    
if __name__ == '__main__': # c'est pour savoir si le program demarre ici ou bien c'est pas le cas
    app = QApplication(argv) #demarrer l'application PyQt
    ex = App() # appler une instance de la class App
    exit(app.exec_()) # la mainloop des GUIs

""" TODO ''''''
TODO_01 : Implement a progress bar XXXXXXX
TODO_02 : MessageBox Or PopUp XXXXXXXX
TODO_03 : when you tap Tab the notabene will make a new \t but i want to go to the frame XXXXXXXXXXXXXX Annulé
TODO_04 : TP ou bien TD ???  XXXXXXXXXXXXXXXXXX
TODO_05 : Année pour le sujet de l'email car on peut refaire l'année XXXXXXXXX
TODO_06 : manage exceptions XXXXXXXXXXXXXX 5/8 erreur capturé
TODO_07 : A propos de l'appli, (how to to send the emails and credits) XXXXXXXX
TODO_08 : Limitations button (how much you can send per day per email provider)/(only 3 sevice providers) XXXXXXX 
"""

#c'est pour les parano 
"""

Je sais que c'est dure de faire confiance a un étudiant, donc je vous invite à créer votre propre fichier .exe !

/1/ aprés la vérification du code; installer l'interpreteur python v3.7 ou plus

/2/aprés l'installation de python , vous irai sur la CMD de windows;(testé seulement sur windows)
    vous installer les modules suivants avec la commande : python -m pip install [le module].

    les module à installer sont :
        -) PyQt5
        -) openpyxl
        -) PyInstaller

/3/ puis vous allez sur le directory de votre nouveau fichier .exe (créez un nouveau dossier) avec : cd [le chemin]

/4/ aprés ça vous taper sur la cmd : PyInstaller --onefile --noconsole [rajouter le chemin de ce fichier .py].
    puis tapez entrer
    
/5/ puis entrez la même commande : PyInstaller --onefile --noconsole [rajouter le chemin vers le fichier créé .spec]

/6/ Enfin vous allez sur le nouveau dossier créé avant, entrez un dossier avec le nom [dist] vous trouvrais un fichier avec l'extension .exe

P.S: N'intoduisez pas les crochets, c'etais juste pour expliquer.
Terminée.
"""
